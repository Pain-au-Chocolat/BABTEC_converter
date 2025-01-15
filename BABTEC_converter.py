import os
import re
import random
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from isofits import *
from License_file import License
from License_file import check_expiry

#Checks if user has valid license key
License()

Number_of_Files = 0
Number_of_Characteristics = 0
Number_of_Errors = 0

#get windows path for this exact file
dir_path = os.getcwd()

#iterate over all Excel files in this path
for filename in os.listdir(dir_path):
    if filename.endswith('.xlsx'):  # Check for Excel files
        file_path = os.path.join(dir_path, filename)  # Create full file path
        workbook = load_workbook(file_path)  # Load the workbook
        sh = workbook["Sheet"]
        Number_of_Files = Number_of_Files + 1

        #get Excel naming for rows and columns. (from 2,5 to B5)
        def xlref(row, column):
            return get_column_letter(column) + str(row)

        #get Position of cell with text "No."
        def cell_no():
            for i in range(1, 1000):
                if sh.cell(i, 1).value == "No.":
                    return [i, 1]
        POS_Cell_No = cell_no()

        # If cell with "No." wasn't found, skip this file
        if not any(char.isdigit() for char in str(POS_Cell_No)):
            print("---------------------------------------------------")
            print("File in this directory:\n" + file_path + "\nIs not supported or created in BABTEC format.")
            print("---------------------------------------------------")
            Number_of_Files = Number_of_Files - 1
            continue

        for i in range(3):
            print("---------------------------------------------------")
        print("File: " + file_path)
        print("No. is at position " + str(POS_Cell_No))

        # get Position of cell with text "Part#"
        def cell_part():
            cell_no_list = []
            for i in range(1, 1000):
                if sh.cell(POS_Cell_No[0], i).value == "Part#\n":
                    cell_no_list.append(POS_Cell_No[0])
                    cell_no_list.append(i)
            return cell_no_list
        POS_Cell_Part = cell_part()
        print("Part# is at position " + str(POS_Cell_Part))

        # get Position of cell with first characteristic
        def cell_first_characteristic():
            for i in range(1,1000):
                finder = str(sh.cell(POS_Cell_No[0]+1, i).value)
                if any(c.isalpha() for c in finder) and finder not in ["None", "KMS"]:
                    return [POS_Cell_No[0]+1, i]
        POS_Cell_First_Characteristic = cell_first_characteristic()
        print("First Characteristic is at position " + str(POS_Cell_First_Characteristic))

        # get Position of cell with last characteristic
        def cell_last_characteristic():
            for i in range(1, 1000):
                finder = str(sh.cell(POS_Cell_First_Characteristic[0] + i, 1).value)
                if finder == "None":
                    return [POS_Cell_First_Characteristic[0] + i - 1, POS_Cell_First_Characteristic[1]]
        POS_Cell_Last_Characteristic = cell_last_characteristic()
        print("Last Characteristic is at position " + str(POS_Cell_Last_Characteristic))

        #Read provided string and find right tolerances to output.
        #Example output : (19.9, 20.1) or (I.O)
        #Example output : (LT, UT) or (text)
        def tolerance_decoder(string):
            #0: Undefined (I.O) Implemented
            #1: Diameter (7.5 , 8.5) Implemented
            #2: Length (7.5 , 8.5) Implemented
            #3: Radius (7.5 , 8.5) Implemented
            #4: Angle (44.0 , 46.0) Implemented
            #5: Chamfer (0.9 , 1.1) Implemented
            #6: Roughness (5.0 , 10.0)

            print("Focus on: " + string)

            if 'µm' in string:
                print("Elox layer found")
                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                tolerance = tolerance.replace("µm", "")
                print("Nominal is: " + nominal)
                print("Tolerance is: " + tolerance)

                if "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                else:
                    if " " in tolerance:
                        tolerance = tolerance.replace("+", "")
                        tol1, tol2 = tolerance.split()
                        if float(tol1) > float(tol2):
                            upper_tolerance = tol1
                            lower_tolerance = tol2
                        else:
                            upper_tolerance = tol2
                            lower_tolerance = tol1
                        print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                        print("Lower tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                        return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                    else:
                        if "-" in tolerance:
                            print("Upper tolerance is: " + str(float(nominal)))
                            print("Lower tolerance is: " + str(float(nominal) + float(tolerance)))
                            return float(nominal) + float(tolerance), float(nominal)
                        else:
                            print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                            print("Lower tolerance is: " + str(float(nominal)))
                            return float(nominal), float(nominal) + float(tolerance)

            if 'Diameter' in string:
                print("Diameter found")
                # ^[^\d]*: This regex pattern matches everything from the start of the string (^), up to the first digit (\d).
                # The [^0-9] part matches any character that is not a digit, and the * means "zero or more" of these non-digit characters.
                # re.sub(r'^[^\d]*', '', s) replaces the matched part (everything before the first digit) with an empty string, effectively removing it.
                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                print("Nominal is: " + nominal)
                print("Tolerance is: " + tolerance)
                if tolerance.isupper():
                    upper_tolerance = isotol('hole', float(nominal), str(tolerance), 'upper') * 0.001 #returns upper tolerance in mm
                    lower_tolerance = isotol('hole', float(nominal), str(tolerance), 'lower') * 0.001 #returns upper tolerance in mm
                    print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(lower_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                elif tolerance.islower():
                    upper_tolerance = isotol('shaft', float(nominal), str(tolerance),'upper') * 0.001  # returns upper tolerance in mm
                    lower_tolerance = isotol('shaft', float(nominal), str(tolerance),'lower') * 0.001  # returns upper tolerance in mm
                    print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(lower_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                elif "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                else:
                    if " " in tolerance:
                        tolerance = tolerance.replace("+", "")
                        tol1, tol2 = tolerance.split()
                        if float(tol1) > float(tol2):
                            upper_tolerance = tol1
                            lower_tolerance = tol2
                        else:
                            upper_tolerance = tol2
                            lower_tolerance = tol1
                        print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                        print("Lower tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                        return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                    else:
                        if "-" in tolerance:
                            print("Upper tolerance is: " + str(float(nominal)))
                            print("Lower tolerance is: " + str(float(nominal) + float(tolerance)))
                            return float(nominal) + float(tolerance), float(nominal)
                        else:
                            print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                            print("Lower tolerance is: " + str(float(nominal)))
                            return float(nominal), float(nominal) + float(tolerance)

            if 'Length' in string:
                print("Length found")
                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                print("Nominal is: " + nominal)
                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                print("Tolerance is: " + tolerance)

                if "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                elif " " in tolerance:
                    tolerance = tolerance.replace("+", "")
                    tol1, tol2 = tolerance.split()
                    if float(tol1) > float(tol2):
                        upper_tolerance = tol1
                        lower_tolerance = tol2
                    else:
                        upper_tolerance = tol2
                        lower_tolerance = tol1
                    print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                else:
                    if "-" in tolerance:
                        print("Upper tolerance is: " + str(float(nominal)))
                        print("Lower tolerance is: " + str((float(nominal) + float(tolerance))))
                        return float(nominal) + float(tolerance), float(nominal)
                    else:
                        print("Upper tolerance is: " + str((float(nominal) + float(tolerance))))
                        print("Lower tolerance is: " + str(float(nominal)))
                        return float(nominal), float(nominal) + float(tolerance)

            if 'Radius' in string:
                print("Radius found")

                if "max" in string:
                    left_text_out = re.sub(r'^[^\d]*', '', string)
                    upper_tolerance = left_text_out.replace(",", ".")
                    lower_tolerance = float(upper_tolerance) * 0.9
                    print("Upper tolerance is: " + str(round(float(upper_tolerance), 3)))
                    print("Lower tolerance is: " + str(round(float(lower_tolerance), 3)))
                    return float(lower_tolerance), float(upper_tolerance)
                if "min" in string:
                    left_text_out = re.sub(r'^[^\d]*', '', string)
                    lower_tolerance = left_text_out.replace(",", ".")
                    upper_tolerance = float(lower_tolerance) * 1.1
                    print("Upper tolerance is: " + str(round(float(upper_tolerance), 3)))
                    print("Lower tolerance is: " + str(round(float(lower_tolerance), 3)))
                    return float(lower_tolerance), float(upper_tolerance)

                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                print("Nominal is: " + nominal)
                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                print("Tolerance is: " + tolerance)

                if "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                elif " " in tolerance:
                    tolerance = tolerance.replace("+", "")
                    tol1, tol2 = tolerance.split()
                    if float(tol1) > float(tol2):
                        upper_tolerance = tol1
                        lower_tolerance = tol2
                    else:
                        upper_tolerance = tol2
                        lower_tolerance = tol1
                    print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                else:
                    if "-" in tolerance:
                        print("Upper tolerance is: " + str(float(nominal)))
                        print("Lower tolerance is: " + str((float(nominal) + float(tolerance))))
                        return float(nominal) + float(tolerance), float(nominal)
                    else:
                        print("Upper tolerance is: " + str((float(nominal) + float(tolerance))))
                        print("Lower tolerance is: " + str(float(nominal)))
                        return float(nominal), float(nominal) + float(tolerance)

            if 'Angle' in string:
                print("Angle found")
                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                left_text_out = left_text_out.replace("°", "")
                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                print("Nominal is: " + nominal)
                if "x" in nominal:
                    nominal = nominal.split('x')[1]
                    lower_tolerance = float(nominal) * 0.85
                    upper_tolerance = float(nominal) * 1.15
                    print("Upper tolerance is: " + str(round(upper_tolerance, 3)))
                    print("Lower tolerance is: " + str(round(lower_tolerance, 3)))
                    return lower_tolerance, upper_tolerance

                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                print("Tolerance is: " + tolerance)

                if "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                elif " " in tolerance:
                    tolerance = tolerance.replace("+", "")
                    tol1, tol2 = tolerance.split()
                    if float(tol1) > float(tol2):
                        upper_tolerance = tol1
                        lower_tolerance = tol2
                    else:
                        upper_tolerance = tol2
                        lower_tolerance = tol1
                    print("Upper tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                else:
                    if "-" in tolerance:
                        print("Upper tolerance is: " + str(float(nominal)))
                        print("Lower tolerance is: " + str((float(nominal) + float(tolerance))))
                        return float(nominal) + float(tolerance), float(nominal)
                    else:
                        print("Upper tolerance is: " + str((float(nominal) + float(tolerance))))
                        print("Lower tolerance is: " + str(float(nominal)))
                        return float(nominal), float(nominal) + float(tolerance)

            if 'Chamfer' in string:
                print("Chamfer found")

                if "max" in string:
                    left_text_out = re.sub(r'^[^\d]*', '', string)
                    upper_tolerance = left_text_out.replace(",", ".")
                    lower_tolerance = float(upper_tolerance) * 0.9
                    print("Upper tolerance is: " + str(round(float(upper_tolerance), 3)))
                    print("Lower tolerance is: " + str(round(float(lower_tolerance), 3)))
                    return float(lower_tolerance), float(upper_tolerance)
                if "min" in string:
                    left_text_out = re.sub(r'^[^\d]*', '', string)
                    lower_tolerance = left_text_out.replace(",", ".")
                    upper_tolerance = float(lower_tolerance) * 1.1
                    print("Upper tolerance is: " + str(round(float(upper_tolerance), 3)))
                    print("Lower tolerance is: " + str(round(float(lower_tolerance), 3)))
                    return float(lower_tolerance), float(upper_tolerance)

                left_text_out = re.sub(r'^[^\d]*', '', string)
                print("Nominal and tolerance is: " + left_text_out)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                if " " not in left_text_out:
                    left_text_out = left_text_out.replace(",", ".")
                    upper_tolerance = float(left_text_out) * 1.1
                    lower_tolerance = float(left_text_out) * 0.9
                    print("Upper tolerance is: " + str(upper_tolerance))
                    print("Lower tolerance is: " + str(lower_tolerance))
                    return float(lower_tolerance), float(upper_tolerance)

                pre_nominal = left_text_out.split(" ", 1)[0]
                nominal = pre_nominal.replace(",", ".")
                print("Nominal is: " + nominal)
                tolerance = left_text_out.split(" ", 1)[1]
                tolerance = tolerance.replace(",", ".")
                print("Tolerance is: " + tolerance)

                if "±" in tolerance:
                    tolerance = tolerance.replace("±", "")
                    print("Upper tolerance is: " + str(float(nominal) + float(tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) - float(tolerance)))
                    return float(nominal) - float(tolerance), float(nominal) + float(tolerance)

                elif " " in tolerance:
                    tolerance = tolerance.replace("+", "")
                    tol1, tol2 = tolerance.split()
                    if float(tol1) > float(tol2):
                        upper_tolerance = tol1
                        lower_tolerance = tol2
                    else:
                        upper_tolerance = tol2
                        lower_tolerance = tol1
                    print("Upper tolerance is: " + str(float(nominal) + float(upper_tolerance)))
                    print("Lower tolerance is: " + str(float(nominal) + float(lower_tolerance)))
                    return float(nominal) + float(lower_tolerance), float(nominal) + float(upper_tolerance)
                else:
                    if "-" in tolerance:
                        print("Upper tolerance is: " + str(float(nominal)))
                        print("Lower tolerance is: " + str((float(nominal) + float(tolerance))))
                        return float(nominal) + float(tolerance), float(nominal)
                    else:
                        print("Upper tolerance is: " + str((float(nominal) + float(tolerance))))
                        print("Lower tolerance is: " + str(float(nominal)))
                        return float(nominal), float(nominal) + float(tolerance)

            if 'Roughness' in string or "roughness" in string:
                print("Roughness found")
                string = string.replace(",", ".")
                left_text_out = re.sub(r'[^\d.]', '', string)
                tolerance = left_text_out.split(" ", 1)[0]
                print("Tolerance is: " + tolerance)
                lower_tolerance = float(tolerance) * 0.1
                upper_tolerance = float(tolerance) * 0.6
                print("Upper tolerance is: " + str(round(upper_tolerance, 3)))
                print("Lower tolerance is: " + str(round(lower_tolerance, 3)))
                return lower_tolerance, upper_tolerance

            GDandT_Symbols = ['Profile', 'Flatness', 'Concentricity', 'Perpendicularity', 'Runout', 'Parallelism',
                        'Circularity', 'Straightness', 'Cylindricity', 'Symmetry', 'Angularity', 'Position']
            if any(keyword in string for keyword in GDandT_Symbols):
                print("GD&T Symbol found")
                string = string.replace(",", ".")
                left_text_out = re.sub(r'[^0-9.]', '', string)
                left_text_out = left_text_out.replace("+ ", "+")
                left_text_out = left_text_out.replace("- ", "-")
                tolerance = left_text_out.split(" ", 1)[0]
                print("Tolerance is: " + tolerance)
                lower_tolerance = float(tolerance) * 0.1
                upper_tolerance = float(tolerance) * 0.6
                print("Upper tolerance is: " + str(round(upper_tolerance, 3)))
                print("Lower tolerance is: " + str(round(lower_tolerance, 3)))
                return lower_tolerance, upper_tolerance


            else:
                print("UNDEFINED / Interpreting as I.O")
                return "I.O"

        # Lowers spread of provided tolerances to be more tight/realistic.
        # This feature won't be implemented if not needed.
        #def tolerance_modifier(lt, ut):
        #    tolerance_spread = round(abs(ut - lt), 3)
        #    tight = round(random.uniform(0.2, 0.4), 3)
        #    tight_tolerance_spread = round(tolerance_spread * tight, 3)
        #    tight_ut = round(ut - tight_tolerance_spread, 3)
        #    tight_lt = round(lt + tight_tolerance_spread, 3)
        #    return tight_lt, tight_ut

        def result_inputer():
            for i in range(POS_Cell_First_Characteristic[0], POS_Cell_Last_Characteristic[0] + 1):
                for j in range(1, len(POS_Cell_Part), 2):
                    print("---------------------------------------------------")
                    print("File: " + file_path)
                    print("Position at: " + xlref(i, POS_Cell_Part[j]))
                    global Number_of_Characteristics
                    Number_of_Characteristics = Number_of_Characteristics + 1
                    try:
                        tol_range = (tolerance_decoder(sh.cell(i, POS_Cell_First_Characteristic[1]).value))
                        if tol_range == "I.O":
                            result = "I.O"
                            sh.cell(i, POS_Cell_Part[j], value=str(result))
                            sh.cell(i, POS_Cell_Part[j]).font = Font(name= "Arial", size=7, color = "000000")
                            print(">>> " + result + " <<<")
                            continue

                        tol1 = float(tol_range[0])
                        tol2 = float(tol_range[1])
                        result = round(random.uniform(tol1, tol2), 3)
                        sh.cell(i, POS_Cell_Part[j], value=str(result))
                        sh.cell(i, POS_Cell_Part[j]).font = Font(name= "Arial", size=7, color="000000")
                        print(">>> " + str(result) + " <<<")

                    except:
                        sh.cell(i, POS_Cell_Part[j], value=str(""))
                        sh.cell(i, POS_Cell_Part[j]).font = Font(name="Arial", size=7, color="000000")
                        global  Number_of_Errors
                        Number_of_Errors = Number_of_Errors + 1
                        print("ERROR")

        result_inputer()

        workbook.save(file_path)
        workbook.close()

for i in range(5):
    print("---------------------------------------------------")
print("Script finished.")

if Number_of_Files == 1:
    print("1 file affected.")
else:
    print(str(Number_of_Files) + " files affected.")

if Number_of_Characteristics == 1:
    print("1 value inputted.")
else:
    print(str(Number_of_Characteristics) + " values inputted.")

if Number_of_Errors == 1:
    print("1 error detected")
else:
    print(str(Number_of_Errors) + " errors detected")

print("---------------------------------------------------")
check_expiry()
print("---------------------------------------------------")
print("*** Program created by unknown... ;) ***")
input()
